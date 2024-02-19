from openpyxl import load_workbook

class XlsxParser:
    def __init__(self) -> None:
        self.__letters = "abcdefghijklmnopqrstuvwxyz"

    def open_book(self, file_path) -> None:
        if file_path:
            self.workbook = load_workbook(f"{file_path}", read_only=True)

    def parse_sheet(self, sheet_name: str, columns_letters: list = None, row_start: int = None) -> dict:
        return {
            sheet_name: self.__parse_rows(
                self.workbook[sheet_name],
                columns_letters, row_start
            )
        }

    def __parse_rows(self, sheet, columns_letters, row_start) -> list:
        rows = []
        
        for row in sheet.iter_rows(
                min_row = row_start if row_start else 1,
                values_only=True
            ):
            rows.append(self.__parse_cells(row, columns_letters))

        return rows
    
    def __parse_cells(self, row, columns_letters) -> list:
        if columns_letters:
            row_cells = []
            for column in columns_letters:
                row_cells.append(row[
                    self.__letters.index(column[-1]) + (len(column)-1)*(len(column)+1)
                ])
        else: return row