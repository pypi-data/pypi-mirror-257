from robot.api.deco import keyword, not_keyword
from openpyxl import Workbook
from datetime import datetime

@keyword("Make Grade Report", types={'answers': list, 'filename': list ,'results': list[list]})
def make_report(answers, filename ,results):
    workbook = Workbook()
    report_file = f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    summarizeSheet = workbook.create_sheet(title=f"Summary")
    header = ['Name'] + [str(i) for i in range(1, len(answers)+1)]+ ['Score']
    summarizeSheet.append(header)

    for (name, result) in zip(filename, results):
        summary = []
        lname = name.split('.')
        lname.pop()
        lname = '_'.join(lname)
        summary.append(lname)

        sheet = workbook.create_sheet(title=f"Result_{lname}")

        # Write header
        header = ['No', 'Answer', 'Result', 'Correct']
        sheet.append(header)

        # Write data rows
        scores = 0
        for(i,ans, res) in zip(range(1, len(answers)+1),answers, result):
            validation = ans == res
            row = [i,ans,res, validation]
            sheet.append(row)
            summary.append(res)
            scores += validation
        sheet.append(['','','Total',scores])
        summary.append(scores)

        summarizeSheet.append(summary)

    # Remove the default sheet created by openpyxl
    workbook.remove(workbook['Sheet'])

    # Save the Excel file
    workbook.save(report_file)

    return report_file

