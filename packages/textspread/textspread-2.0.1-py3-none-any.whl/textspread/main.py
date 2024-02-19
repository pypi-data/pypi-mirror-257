#!/usr/bin/env python3

from argparse import ArgumentParser
from os import path
from openpyxl import load_workbook, Workbook
from datetime import datetime

EXIT_SUCCESS = 0
EXIT_ERROR_FILE_NOT_FOUND = 1


def cli():
    parser = ArgumentParser(prog="textspread",
                            description="Textspread takes in a .txt file and \
                                    appends its contents into a spreadsheet",
                            epilog="Author: [t0mri](https://github.com/t0mri)")

    parser.add_argument("files", nargs="*",
                        help="Files to be appended into the spreadsheet")
    parser.add_argument("-c", "--columns", required=True,
                        type=int, help="Target columns count")
    parser.add_argument("-a", "--append",
                        help="Append it to the given file")
    parser.add_argument("-f", "--format", default="xlsx",
                        choices=["xls", "xlsx", "ods"], help="Output format")

    args = parser.parse_args()

    XL_FILE = args.append

    if XL_FILE and not path.isfile(XL_FILE):
        print(XL_FILE, "not found!")
        exit(EXIT_ERROR_FILE_NOT_FOUND)

    wb = Workbook()
    if XL_FILE:
        wb = load_workbook(XL_FILE)
    ws = wb.active

    for file in args.files:
        if not path.isfile(file):
            print(file, "not found!")
            exit(EXIT_ERROR_FILE_NOT_FOUND)

        file = open(file, "r")
        lineTokens = file.readlines()

        for i in range(0, len(lineTokens), args.columns):
            ws.append(lineTokens[i:args.columns+args.columns*i])

    if XL_FILE:
        wb.save(XL_FILE)
    else:
        wb.save(datetime.now().strftime("%Y%m%d%H%M%S")+"."+args.format)
    exit(EXIT_SUCCESS)


cli()
