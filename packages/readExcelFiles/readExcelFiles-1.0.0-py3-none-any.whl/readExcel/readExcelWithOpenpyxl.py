from openpyxl import load_workbook, Workbook

"""
   Read data from an Excel file and return it as a list of rows.

   Args:
       file_path (str): The file path of the Excel file.
       sheet_name (str): The name of the worksheet. Default is 'Sheet1'.

   Returns:
       list: A list containing rows of data from the specified worksheet.
   """


def readExcelWithOpenpyxl(file_path, sheet_name='Sheet1'):
    global workbook
    try:
        # Load Excel workbook
        workbook: Workbook = load_workbook(file_path)

        # Access a specific worksheet
        worksheet = workbook[sheet_name]

        # Accessing data from the worksheet
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)

        return data

    finally:
        # Make sure to close the workbook
        workbook.close()
