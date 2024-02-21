from openpyxl import load_workbook, Workbook

import warnings

# Disable all warnings
warnings.filterwarnings("ignore")


def readExcelWithOpenpyxl(excelFilePath, sheetName='Sheet1'):
    """
       Read data from an Excel file and return it as a list of rows.

       Args:
           excelFilePath (str): The file path of the Excel file.
           sheetName (str): The name of the worksheet. Default is 'Sheet1'.

       Returns:
           list: A list containing rows of data from the specified worksheet.
       """
    workbook = None
    try:
        # Load Excel workbook
        workbook: Workbook = load_workbook(excelFilePath)

        # Access a specific worksheet
        worksheet = workbook[sheetName]

        # Accessing data from the worksheet
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)

        return data

    except Exception as e:
        # Handle any exceptions that occur during execution
        print(f"An error occurred: {e}")
        return None

    finally:
        # Make sure to close the workbook
        if workbook is not None:
            workbook.close()
